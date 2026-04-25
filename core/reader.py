"""Workbook readers with a shared internal representation."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import TypeAlias
from zipfile import BadZipFile

import xlrd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

CellValue: TypeAlias = str | int | float | bool | None
RowData: TypeAlias = tuple[CellValue, ...]


@dataclass(frozen=True)
class SheetData:
    name: str
    rows: tuple[RowData, ...]


@dataclass(frozen=True)
class WorkbookData:
    path: Path
    sheets: tuple[SheetData, ...]


class ReaderError(Exception):
    """Raised when a workbook cannot be read into the internal representation."""


def read_workbook(path: str | Path) -> WorkbookData:
    workbook_path = Path(path)
    _validate_readable_file(workbook_path)

    suffix = workbook_path.suffix.lower()
    if suffix == ".xlsx":
        workbook = _read_xlsx(workbook_path)
    elif suffix == ".xls":
        workbook = _read_xls(workbook_path)
    else:
        raise ReaderError(
            f"Unsupported workbook type for '{workbook_path}': expected .xls or .xlsx"
        )

    if not workbook.sheets:
        raise ReaderError(f"Workbook '{workbook_path}' contains no sheets")

    if not any(sheet.rows for sheet in workbook.sheets):
        raise ReaderError(f"Workbook '{workbook_path}' contains no data rows")

    return workbook


def _validate_readable_file(path: Path) -> None:
    try:
        stat_result = path.stat()
    except FileNotFoundError as exc:
        raise ReaderError(f"Workbook '{path}' does not exist") from exc
    except PermissionError as exc:
        raise ReaderError(f"Workbook '{path}' is not readable or is locked") from exc

    if not path.is_file():
        raise ReaderError(f"Workbook '{path}' is not a file")

    if stat_result.st_size == 0:
        raise ReaderError(f"Workbook '{path}' is empty")


def _read_xlsx(path: Path) -> WorkbookData:
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except PermissionError as exc:
        raise ReaderError(f"Workbook '{path}' is not readable or is locked") from exc
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise ReaderError(f"Workbook '{path}' could not be read as .xlsx: {exc}") from exc

    try:
        sheets = tuple(
            SheetData(
                name=worksheet.title,
                rows=tuple(tuple(row) for row in worksheet.iter_rows(values_only=True)),
            )
            for worksheet in workbook.worksheets
        )
    finally:
        workbook.close()

    return WorkbookData(path=path, sheets=sheets)


def _read_xls(path: Path) -> WorkbookData:
    try:
        workbook = xlrd.open_workbook(filename=str(path))
    except PermissionError as exc:
        raise ReaderError(f"Workbook '{path}' is not readable or is locked") from exc
    except (xlrd.XLRDError, OSError) as exc:
        raise ReaderError(f"Workbook '{path}' could not be read as .xls: {exc}") from exc

    sheets = tuple(
        SheetData(
            name=sheet.name,
            rows=tuple(
                tuple(sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols))
                for row_index in range(sheet.nrows)
            ),
        )
        for sheet in workbook.sheets()
    )

    return WorkbookData(path=path, sheets=sheets)
