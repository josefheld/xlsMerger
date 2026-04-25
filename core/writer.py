"""Workbook writer for the shared internal representation."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from core.reader import WorkbookData


class WriterError(Exception):
    """Raised when workbook data cannot be written."""


def write_workbook(workbook: WorkbookData, path: str | Path) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    target = Workbook()
    default_sheet = target.active
    target.remove(default_sheet)

    for sheet in workbook.sheets:
        worksheet = target.create_sheet(title=sheet.name)
        for row in sheet.rows:
            worksheet.append(row)

    if not workbook.sheets:
        target.create_sheet(title="Sheet1")

    try:
        target.save(output_path)
    except OSError as exc:
        raise WriterError(f"Workbook '{output_path}' could not be written: {exc}") from exc

    return output_path
