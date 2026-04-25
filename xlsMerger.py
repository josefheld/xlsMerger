import os

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def open_xls_as_xlsx(filename: str) -> Workbook:
    xls_book = xlrd.open_workbook(filename=filename)
    workbook = Workbook()

    for sheet_index in range(0, xls_book.nsheets):
        xls_sheet = xls_book.sheet_by_index(sheet_index)
        sheet = workbook.active if sheet_index == 0 else workbook.create_sheet()
        sheet.title = xls_sheet.name

        for row in range(0, xls_sheet.nrows):
            for col in range(0, xls_sheet.ncols):
                sheet.cell(row=row + 1, column=col + 1).value = xls_sheet.cell_value(row, col)
    return workbook


def main() -> None:
    combined_wb = Workbook()
    del combined_wb['Sheet']
    mypath = './Originale'
    filenames = [f for f in os.listdir(mypath) if os.path.isfile(os.path.join(mypath, f))]

    file_count = 0
    last_row_num: dict[str, int] = {}
    for filename in filenames:
        if not filename.endswith('xls') or filename.endswith('xlsx'):
            continue
        fullname = os.path.join(mypath, filename)

        converted_original_book = open_xls_as_xlsx(fullname)
        converted_original_book.Template = False

        for sheetname in converted_original_book.sheetnames:
            if sheetname not in last_row_num:
                last_row_num[sheetname] = 0

        for sheetname in converted_original_book.sheetnames:
            if sheetname not in combined_wb.sheetnames:
                combined_wb.create_sheet(title=sheetname)
            old_sheet = converted_original_book[sheetname]
            new_sheet = combined_wb[sheetname]

            start_row = 1 if file_count == 0 else 2
            rows_copied = 0

            for row_num in range(start_row, old_sheet.max_row + 1):
                first_cell_value = old_sheet.cell(row=row_num, column=1).value
                if first_cell_value is None:
                    break

                for col_num in range(1, old_sheet.max_column + 1):
                    cell_value = old_sheet.cell(row=row_num, column=col_num).value
                    if cell_value is None:
                        break
                    ft1 = Font(name='Arial', size=10, color='FF000000')
                    if row_num == 1:
                        ft1 = Font(name='Arial', size=10, color='FFFFFFFF')
                        fill = PatternFill(
                            fill_type='lightTrellis',
                            start_color='FF1C0082',
                            end_color='FF1C0082',
                        )
                    else:
                        ft1 = Font(name='Arial', size=10, color='FF000000')
                        fill = PatternFill(
                            fill_type='lightTrellis',
                            start_color='FFFFFFFF',
                            end_color='FFFFFFFF',
                        )
                    new_cell = new_sheet.cell(row=row_num + last_row_num[sheetname], column=col_num)
                    new_cell.value = cell_value
                    new_cell.font = ft1
                    new_cell.fill = fill

                rows_copied += 1

            last_row_num[sheetname] += rows_copied
        file_count += 1

    dest_filename = 'Merged.xlsx'

    combined_wb.save(filename=dest_filename)


if __name__ == '__main__':
    main()
