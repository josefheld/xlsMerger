from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from os import listdir
from os.path import isfile, join
import os
import xlrd
from openpyxl.workbook import Workbook as openpyxlWorkbook
from openpyxl.styles import Font, Color, PatternFill

def open_xls_as_xlsx(filename):
    xlsBook = xlrd.open_workbook(filename=filename)
    workbook = openpyxlWorkbook()

    for i in xrange(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in xrange(0, xlsSheet.nrows):
            for col in xrange(0, xlsSheet.ncols):
                sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)
    return workbook

def main():
    combined_wb = Workbook()
    del combined_wb['Sheet']
    mypath = './Originale'
    filenames = [f for f in listdir(mypath) if isfile(join(mypath, f))]

    file_count = 0
    last_row_num = {}
    for filename in filenames:
        fullname = os.path.join(mypath, filename)

        converted_original_book = open_xls_as_xlsx(fullname)
        converted_original_book.Template = False

        for sheetname in converted_original_book.sheetnames:
            if sheetname not in last_row_num.iterkeys():
                last_row_num[sheetname] = 0

        for sheetname in converted_original_book.sheetnames:
            if not combined_wb.__contains__(sheetname):
               combined_wb.create_sheet(title=sheetname)
            old_sheet = converted_original_book[sheetname]
            new_sheet = combined_wb[sheetname]


            finished_row = False

            start_row = 1 if file_count == 0 else 2

            for row_num in range(start_row, 1000):

                emptyCellCheck_1 = old_sheet.cell(row=row_num, column=1).value
                if emptyCellCheck_1 is None:
                    break;

                for col_num in range(1, 100):
                    emptyCellCheck_2 = old_sheet.cell(row=row_num, column=col_num).value
                    if emptyCellCheck_2 is None:
                        finished_row = True
                        break
                    ft1 = Font(name='Arial', size=10, color='FF000000')
                    if row_num == 1:
                        ft1 = Font(name='Arial', size=10, color='FFFFFFFF')
                        fill = PatternFill(fill_type='lightTrellis', start_color = 'FF1C0082', end_color = 'FF1C0082')
                    else:
                        ft1 = Font(name='Arial', size=10, color='FF000000')
                        fill = PatternFill(fill_type='lightTrellis', start_color = 'FFFFFFFF', end_color = 'FFFFFFFF')
                    new_cell = new_sheet.cell(row=row_num+last_row_num[sheetname], column=col_num)
                    new_cell.value = old_sheet.cell(row=row_num, column=col_num).value
                    new_cell.font = ft1
                    new_cell.fill = fill


            last_row_num[sheetname] = last_row_num[sheetname] + row_num-2
        file_count = file_count + 1
                #if finished_row == True:
                #    break



    dest_filename = 'Zusammengefasste_Dateien.xlsx'

    combined_wb.save(filename=dest_filename)


if __name__ == '__main__': main()