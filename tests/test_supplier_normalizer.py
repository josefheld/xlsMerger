import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from cli import main as cli_main
from core.reader import SheetData, WorkbookData
from profiles import ProfileConfig
from profiles.supplier_normalizer import TARGET_SCHEMA, SupplierNormalizerProfile


def workbook(rows: list[tuple[object, ...]], sheet_name: str = "Invoices") -> WorkbookData:
    return WorkbookData(
        path=Path("supplier.xlsx"),
        sheets=(SheetData(name=sheet_name, rows=tuple(rows)),),
    )


def write_xlsx(path: Path, rows: list[list[object]]) -> None:
    source = Workbook()
    worksheet = source.active
    worksheet.title = "Invoices"
    for row in rows:
        worksheet.append(row)
    source.save(path)


def default_config(**options: object) -> ProfileConfig:
    return ProfileConfig(mode="supplier_normalizer", options=options)


def test_supplier_normalizer_maps_different_layouts_to_same_schema() -> None:
    profile = SupplierNormalizerProfile()
    layout_a = workbook(
        [
            ("Vendor", "Invoice No", "PO Number", "Date", "Total", "Currency"),
            ("ACME GmbH", "INV-1", "PO-1", "2026-02-01", 12.5, "eur"),
        ]
    )
    layout_b = workbook(
        [
            ("Lieferant", "Rechnungsnummer", "Bestellnummer", "Betrag", "Waehrung"),
            ("Globex", "INV-2", "PO-2", "19.95", "usd"),
        ]
    )

    transformed_a = profile.transform(layout_a, default_config())
    transformed_b = profile.transform(layout_b, default_config())

    assert transformed_a.sheets[0].rows == (
        TARGET_SCHEMA,
        ("ACME GmbH", "INV-1", "PO-1", "2026-02-01", 12.5, "EUR"),
    )
    assert transformed_b.sheets[0].rows == (
        TARGET_SCHEMA,
        ("Globex", "INV-2", "PO-2", "", 19.95, "USD"),
    )


def test_supplier_normalizer_uses_configured_column_synonyms_and_supplier_name() -> None:
    profile = SupplierNormalizerProfile()
    source = workbook(
        [
            ("Bill ID", "Order Ref", "Gross"),
            ("INV-1", "PO-1", "12.50"),
        ]
    )
    config = default_config(
        supplier_name="Acme Raw",
        supplier_name_map={"Acme Raw": "ACME GmbH"},
        default_currency="chf",
        column_synonyms={
            "invoice_id": ["Bill ID"],
            "order_id": ["Order Ref"],
            "amount": ["Gross"],
        },
    )

    issues = profile.validate(source, config)
    transformed = profile.transform(source, config)

    assert issues == ()
    assert transformed.sheets[0].rows == (
        TARGET_SCHEMA,
        ("ACME GmbH", "INV-1", "PO-1", "", 12.5, "CHF"),
    )


def test_supplier_normalizer_reports_missing_required_columns_precisely() -> None:
    profile = SupplierNormalizerProfile()
    source = workbook(
        [
            ("Vendor", "Invoice No"),
            ("ACME GmbH", "INV-1"),
        ]
    )

    issues = profile.validate(source, default_config())

    assert len(issues) == 1
    assert issues[0].code == "missing_required_column"
    assert issues[0].sheet == "Invoices"
    assert "missing required columns: order_id, amount" in issues[0].message


def test_supplier_normalizer_marks_duplicates_in_report(tmp_path: Path) -> None:
    input_path = tmp_path / "supplier.xlsx"
    report_path = tmp_path / "report.json"
    write_xlsx(
        input_path,
        [
            ["Vendor", "Invoice No", "PO Number", "Total"],
            ["ACME GmbH", "INV-1", "PO-1", 12.5],
            ["ACME GmbH", "INV-1", "PO-1", 12.5],
        ],
    )

    exit_code = cli_main.main(
        [
            "run",
            str(input_path),
            "--mode",
            "supplier_normalizer",
            "--report",
            str(report_path),
        ]
    )

    payload = json.loads(report_path.read_text(encoding="utf-8"))
    assert exit_code == 0
    assert payload["warnings"][0]["code"] == "duplicate_invoice_order"
    assert "duplicates invoice_id 'INV-1' and order_id 'PO-1'" in payload["warnings"][0]["message"]


def test_supplier_normalizer_cli_output_file_matches_target_schema(tmp_path: Path) -> None:
    input_path = tmp_path / "supplier.xlsx"
    output_path = tmp_path / "supplier-output.xlsx"
    report_path = tmp_path / "report.json"
    write_xlsx(
        input_path,
        [
            ["Lieferant", "Rechnungsnummer", "Bestellnummer", "Betrag", "Waehrung"],
            ["Globex", "INV-2", "PO-2", 19.95, "usd"],
        ],
    )

    exit_code = cli_main.main(
        [
            "run",
            str(input_path),
            "--mode",
            "supplier_normalizer",
            "--output",
            str(output_path),
            "--report",
            str(report_path),
        ]
    )

    output = load_workbook(output_path, data_only=True)
    rows = list(output["Invoices"].iter_rows(values_only=True))

    assert exit_code == 0
    assert rows == [
        TARGET_SCHEMA,
        ("Globex", "INV-2", "PO-2", None, 19.95, "USD"),
    ]
