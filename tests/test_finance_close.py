import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from cli import main as cli_main
from core.reader import SheetData, WorkbookData
from profiles import ProfileConfig
from profiles.finance_close import TARGET_SCHEMA, FinanceCloseProfile


def workbook(rows: list[tuple[object, ...]], sheet_name: str = "Ledger") -> WorkbookData:
    return WorkbookData(
        path=Path("finance.xlsx"),
        sheets=(SheetData(name=sheet_name, rows=tuple(rows)),),
    )


def write_xlsx(path: Path, rows: list[list[object]]) -> None:
    source = Workbook()
    worksheet = source.active
    worksheet.title = "Ledger"
    for row in rows:
        worksheet.append(row)
    source.save(path)


def default_config() -> ProfileConfig:
    return ProfileConfig(mode="finance_close", options={})


def test_finance_close_reports_missing_required_columns_precisely() -> None:
    profile = FinanceCloseProfile()
    source = workbook(
        [
            ("date", "account", "debit"),
            ("2026-01-31", "1000", 10),
        ]
    )

    issues = profile.validate(source, default_config())

    assert len(issues) == 1
    assert issues[0].code == "missing_required_column"
    assert issues[0].sheet == "Ledger"
    assert "missing required columns: credit, balance" in issues[0].message


def test_finance_close_validates_date_and_number_columns() -> None:
    profile = FinanceCloseProfile()
    source = workbook(
        [
            ("date", "account", "debit", "credit", "balance"),
            ("31/01/2026", "1000", "not-a-number", 0, 0),
        ]
    )

    issues = profile.validate(source, default_config())

    assert [issue.code for issue in issues] == ["invalid_column_type", "invalid_column_type"]
    assert "row 2 column 'date' must be a date in YYYY-MM-DD format" in issues[0].message
    assert "row 2 column 'debit' must be a number" in issues[1].message


def test_finance_close_checks_debit_credit_against_closing_balance() -> None:
    profile = FinanceCloseProfile()
    source = workbook(
        [
            ("date", "account", "debit", "credit", "balance"),
            ("2026-01-31", "1000", 10, 3, 99),
        ]
    )

    issues = profile.validate(source, default_config())

    assert len(issues) == 1
    assert issues[0].code == "balance_mismatch"
    assert "has debit-credit total 7 but closing balance 99" in issues[0].message


def test_finance_close_transform_outputs_target_schema() -> None:
    profile = FinanceCloseProfile()
    source = workbook(
        [
            ("credit", "date", "balance", "account", "description", "debit"),
            (3, "2026-01-31", 7, "1000", "Revenue close", 10),
        ]
    )

    transformed = profile.transform(source, default_config())

    assert transformed.sheets[0].rows == (
        TARGET_SCHEMA,
        ("2026-01-31", "1000", "Revenue close", 10.0, 3.0, 7.0),
    )


def test_finance_close_cli_output_file_matches_target_schema(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "finance-output.xlsx"
    report_path = tmp_path / "report.json"
    write_xlsx(
        input_path,
        [
            ["credit", "date", "balance", "account", "description", "debit"],
            [3, "2026-01-31", 7, "1000", "Revenue close", 10],
        ],
    )

    exit_code = cli_main.main(
        [
            "run",
            str(input_path),
            "--mode",
            "finance_close",
            "--output",
            str(output_path),
            "--report",
            str(report_path),
        ]
    )

    output = load_workbook(output_path, data_only=True)
    rows = list(output["Ledger"].iter_rows(values_only=True))
    report = json.loads(report_path.read_text(encoding="utf-8"))

    assert exit_code == 0
    assert rows == [
        TARGET_SCHEMA,
        ("2026-01-31", "1000", "Revenue close", 10, 3, 7),
    ]
    assert report["errors"] == []
