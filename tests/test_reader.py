from pathlib import Path

import pytest
import xlwt
from openpyxl import Workbook

from core import reader
from core.reader import ReaderError, read_workbook


def write_xlsx(path: Path, rows: list[list[object]], sheet_name: str = "Data") -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    for row in rows:
        worksheet.append(row)
    workbook.save(path)


def write_xls(path: Path, rows: list[list[object]], sheet_name: str = "Data") -> None:
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet(sheet_name)
    for row_index, row in enumerate(rows):
        for column_index, value in enumerate(row):
            worksheet.write(row_index, column_index, value)
    workbook.save(path)


def test_read_xlsx_returns_common_representation(tmp_path: Path) -> None:
    path = tmp_path / "input.xlsx"
    write_xlsx(path, [["name", "amount"], ["A", 12]])

    workbook = read_workbook(path)

    assert workbook.path == path
    assert workbook.sheets[0].name == "Data"
    assert workbook.sheets[0].rows == (("name", "amount"), ("A", 12))


def test_read_xls_returns_common_representation(tmp_path: Path) -> None:
    path = tmp_path / "input.xls"
    write_xls(path, [["name", "amount"], ["A", 12]])

    workbook = read_workbook(path)

    assert workbook.path == path
    assert workbook.sheets[0].name == "Data"
    assert workbook.sheets[0].rows == (("name", "amount"), ("A", 12.0))


def test_missing_workbook_has_clear_error(tmp_path: Path) -> None:
    path = tmp_path / "missing.xlsx"

    with pytest.raises(ReaderError, match="does not exist"):
        read_workbook(path)


def test_unsupported_extension_has_clear_error(tmp_path: Path) -> None:
    path = tmp_path / "input.csv"
    path.write_text("name,amount\nA,12\n")

    with pytest.raises(ReaderError, match="expected .xls or .xlsx"):
        read_workbook(path)


def test_empty_file_has_clear_error(tmp_path: Path) -> None:
    path = tmp_path / "empty.xlsx"
    path.touch()

    with pytest.raises(ReaderError, match="is empty"):
        read_workbook(path)


def test_empty_workbook_has_clear_error(tmp_path: Path) -> None:
    path = tmp_path / "empty-workbook.xlsx"
    write_xlsx(path, [])

    with pytest.raises(ReaderError, match="contains no data rows"):
        read_workbook(path)


def test_corrupt_xlsx_has_clear_error(tmp_path: Path) -> None:
    path = tmp_path / "corrupt.xlsx"
    path.write_bytes(b"not an excel workbook")

    with pytest.raises(ReaderError, match="could not be read as .xlsx"):
        read_workbook(path)


def test_locked_workbook_has_clear_error(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    path = tmp_path / "locked.xlsx"
    write_xlsx(path, [["name"]])

    def raise_permission_error(*args: object, **kwargs: object) -> None:
        raise PermissionError("locked")

    monkeypatch.setattr(reader, "load_workbook", raise_permission_error)

    with pytest.raises(ReaderError, match="not readable or is locked"):
        read_workbook(path)
