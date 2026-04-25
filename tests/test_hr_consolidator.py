import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from cli import main as cli_main
from core.reader import SheetData, WorkbookData
from profiles import ProfileConfig
from profiles.hr_consolidator import TARGET_SCHEMA, HrConsolidatorProfile


def workbook(rows: list[tuple[object, ...]], sheet_name: str = "Employees") -> WorkbookData:
    return WorkbookData(
        path=Path("hr.xlsx"),
        sheets=(SheetData(name=sheet_name, rows=tuple(rows)),),
    )


def write_xlsx(path: Path, rows: list[list[object]]) -> None:
    source = Workbook()
    worksheet = source.active
    worksheet.title = "Employees"
    for row in rows:
        worksheet.append(row)
    source.save(path)


def default_config(**options: object) -> ProfileConfig:
    return ProfileConfig(mode="hr_consolidator", options=options)


def test_hr_consolidator_masks_default_pii_fields() -> None:
    profile = HrConsolidatorProfile()
    source = workbook(
        [
            ("employee_id", "first_name", "last_name", "email", "department", "hire_date"),
            ("EMP-001", "Ada", "Lovelace", "ada@example.com", "Engineering", "2026-01-01"),
        ]
    )

    transformed = profile.transform(source, default_config())

    assert transformed.sheets[0].rows == (
        TARGET_SCHEMA,
        (
            "EMP-001",
            "***MASKED***",
            "***MASKED***",
            "***MASKED***",
            "Engineering",
            "2026-01-01",
            "",
        ),
    )


def test_hr_consolidator_respects_configured_mask_fields_and_token() -> None:
    profile = HrConsolidatorProfile()
    source = workbook(
        [
            ("Personalnummer", "Vorname", "Nachname", "Mail", "Eintrittsdatum"),
            ("E123", "Ada", "Lovelace", "ada@example.com", "2026-01-01"),
        ]
    )

    transformed = profile.transform(
        source,
        default_config(mask_fields=["email"], mask_token="[redacted]"),
    )

    assert transformed.sheets[0].rows[1] == (
        "E123",
        "Ada",
        "Lovelace",
        "[redacted]",
        "",
        "2026-01-01",
        "",
    )


def test_hr_consolidator_reports_invalid_id_and_dates() -> None:
    profile = HrConsolidatorProfile()
    source = workbook(
        [
            ("employee_id", "first_name", "last_name", "hire_date", "termination_date"),
            ("x", "Ada", "Lovelace", "01/01/2026", "not-a-date"),
        ]
    )

    issues = profile.validate(source, default_config())

    assert [issue.code for issue in issues] == [
        "invalid_hr_record",
        "invalid_hr_record",
        "invalid_hr_record",
    ]
    assert "row 2 column 'employee_id' must be 3-32 characters" in issues[0].message
    assert "row 2 column 'hire_date' must be a date in YYYY-MM-DD format" in issues[1].message
    assert (
        "row 2 column 'termination_date' must be a date in YYYY-MM-DD format"
        in issues[2].message
    )


def test_hr_consolidator_reports_missing_required_columns_precisely() -> None:
    profile = HrConsolidatorProfile()
    source = workbook(
        [
            ("employee_id", "first_name"),
            ("EMP-001", "Ada"),
        ]
    )

    issues = profile.validate(source, default_config())

    assert len(issues) == 1
    assert issues[0].code == "missing_required_column"
    assert issues[0].sheet == "Employees"
    assert "missing required columns: last_name, hire_date" in issues[0].message


def test_hr_consolidator_cli_output_masks_pii_reliably(tmp_path: Path) -> None:
    input_path = tmp_path / "hr.xlsx"
    output_path = tmp_path / "hr-output.xlsx"
    report_path = tmp_path / "report.json"
    write_xlsx(
        input_path,
        [
            ["employee_id", "first_name", "last_name", "email", "department", "hire_date"],
            ["EMP-001", "Ada", "Lovelace", "ada@example.com", "Engineering", "2026-01-01"],
        ],
    )

    exit_code = cli_main.main(
        [
            "run",
            str(input_path),
            "--mode",
            "hr_consolidator",
            "--output",
            str(output_path),
            "--report",
            str(report_path),
        ]
    )

    output = load_workbook(output_path, data_only=True)
    rows = list(output["Employees"].iter_rows(values_only=True))
    report = json.loads(report_path.read_text(encoding="utf-8"))

    assert exit_code == 0
    assert rows == [
        TARGET_SCHEMA,
        (
            "EMP-001",
            "***MASKED***",
            "***MASKED***",
            "***MASKED***",
            "Engineering",
            "2026-01-01",
            None,
        ),
    ]
    assert "Ada" not in str(rows)
    assert "Lovelace" not in str(rows)
    assert "ada@example.com" not in str(rows)
    assert report["errors"] == []


def test_hr_consolidator_cli_reports_invalid_records(tmp_path: Path) -> None:
    input_path = tmp_path / "hr.xlsx"
    report_path = tmp_path / "report.json"
    write_xlsx(
        input_path,
        [
            ["employee_id", "first_name", "last_name", "hire_date"],
            ["x", "Ada", "Lovelace", "bad-date"],
        ],
    )

    exit_code = cli_main.main(
        ["run", str(input_path), "--mode", "hr_consolidator", "--report", str(report_path)]
    )

    payload = json.loads(report_path.read_text(encoding="utf-8"))
    assert exit_code == 1
    assert [error["code"] for error in payload["errors"]] == [
        "invalid_hr_record",
        "invalid_hr_record",
    ]
    assert "row 2 column 'employee_id'" in payload["errors"][0]["message"]
    assert "row 2 column 'hire_date'" in payload["errors"][1]["message"]
